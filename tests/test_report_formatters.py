"""Unit tests for clarinet.utils.report_formatters."""

import io

from openpyxl import load_workbook

from clarinet.utils.report_formatters import to_csv, to_xlsx


def test_csv_has_header_and_rows() -> None:
    buf = to_csv(["id", "name"], [(1, "alpha"), (2, "beta")])
    text = buf.getvalue().decode("utf-8-sig")
    lines = text.splitlines()
    assert lines[0] == "id,name"
    assert lines[1] == "1,alpha"
    assert lines[2] == "2,beta"


def test_csv_has_utf8_bom_for_excel() -> None:
    buf = to_csv(["x"], [(1,)])
    raw = buf.getvalue()
    assert raw.startswith(b"\xef\xbb\xbf"), "Excel needs UTF-8 BOM for Cyrillic"


def test_csv_handles_unicode() -> None:
    buf = to_csv(["имя"], [("Тест",), ("Алиса",)])
    text = buf.getvalue().decode("utf-8-sig")
    assert "имя" in text
    assert "Тест" in text


def test_csv_replaces_none_with_empty() -> None:
    buf = to_csv(["a", "b"], [(1, None), (None, 2)])
    text = buf.getvalue().decode("utf-8-sig")
    lines = text.splitlines()
    assert lines[1] == "1,"
    assert lines[2] == ",2"


def test_csv_uses_crlf_line_terminator() -> None:
    buf = to_csv(["x"], [(1,), (2,)])
    raw = buf.getvalue()
    assert b"\r\n" in raw


def test_csv_empty_rows_keeps_header() -> None:
    buf = to_csv(["id"], [])
    text = buf.getvalue().decode("utf-8-sig")
    assert text.splitlines() == ["id"]


def test_xlsx_round_trip() -> None:
    buf = to_xlsx(["id", "name"], [(1, "alpha"), (2, "beta")])
    buf.seek(0)
    wb = load_workbook(io.BytesIO(buf.read()), read_only=True)
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("id", "name")
    assert rows[1] == (1, "alpha")
    assert rows[2] == (2, "beta")


def test_xlsx_handles_unicode_and_none() -> None:
    buf = to_xlsx(["имя", "value"], [("Тест", None), ("Алиса", 42)])
    wb = load_workbook(io.BytesIO(buf.getvalue()), read_only=True)
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    # Read-only mode trims trailing None cells from each row, so compare prefixes.
    assert rows[0][:2] == ("имя", "value")
    assert rows[1][:1] == ("Тест",)
    assert rows[2][:2] == ("Алиса", 42)


def test_xlsx_empty_rows_keeps_header() -> None:
    buf = to_xlsx(["only_col"], [])
    wb = load_workbook(io.BytesIO(buf.getvalue()), read_only=True)
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    assert rows == [("only_col",)]
