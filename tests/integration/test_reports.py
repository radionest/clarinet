"""Integration tests for the custom SQL reports endpoints."""

import io
from collections.abc import AsyncGenerator

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook

from clarinet.api.app import app
from clarinet.api.dependencies import get_report_registry
from clarinet.models.report import ReportTemplate
from clarinet.services.report_service import ReportRegistry
from tests.conftest import create_authenticated_client, create_mock_superuser
from tests.utils.urls import ADMIN_REPORTS

pytestmark = pytest.mark.asyncio


def _make_registry() -> ReportRegistry:
    """Build a registry with table-free SELECTs.

    The repository uses ``db_manager`` (not the per-test SQLAlchemy engine),
    so any query that references project tables would fail with ``no such
    table``. Restrict the integration suite to table-free SELECTs and rely on
    the unit tests for parser / formatter coverage.
    """
    return ReportRegistry(
        [
            (
                ReportTemplate(
                    name="constant_one",
                    title="One row",
                    description="Returns a single literal row",
                ),
                "SELECT 1 AS value",
            ),
            (
                ReportTemplate(name="two_columns", title="Two columns", description=""),
                "SELECT 1 AS id, 'alpha' AS name",
            ),
        ]
    )


@pytest_asyncio.fixture
async def report_client(test_session, test_settings) -> AsyncGenerator[AsyncClient]:
    """Authenticated superuser client with a fixed report registry."""
    mock_user = await create_mock_superuser(test_session, email="reports@test.com")
    registry = _make_registry()
    app.dependency_overrides[get_report_registry] = lambda: registry

    async for ac in create_authenticated_client(mock_user, test_session, test_settings):
        yield ac

    app.dependency_overrides.pop(get_report_registry, None)


@pytest_asyncio.fixture
async def unauth_client(test_session, test_settings) -> AsyncGenerator[AsyncClient]:
    """Unauthenticated client for 401 checks."""
    from clarinet.utils.database import get_async_session

    async def override_get_session():
        yield test_session

    app.dependency_overrides[get_async_session] = override_get_session
    app.dependency_overrides[get_report_registry] = lambda: _make_registry()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        yield ac

    app.dependency_overrides.clear()


async def test_list_reports_returns_templates(report_client: AsyncClient) -> None:
    resp = await report_client.get(ADMIN_REPORTS)
    assert resp.status_code == 200
    items = resp.json()
    names = [item["name"] for item in items]
    assert "constant_one" in names
    assert "two_columns" in names
    one = next(i for i in items if i["name"] == "constant_one")
    assert one["title"] == "One row"
    assert one["description"] == "Returns a single literal row"


async def test_list_reports_requires_auth(unauth_client: AsyncClient) -> None:
    resp = await unauth_client.get(ADMIN_REPORTS)
    assert resp.status_code == 401


async def test_download_csv(report_client: AsyncClient) -> None:
    resp = await report_client.get(f"{ADMIN_REPORTS}/constant_one/download")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")
    assert "constant_one_" in resp.headers["content-disposition"]
    assert ".csv" in resp.headers["content-disposition"]
    text = resp.content.decode("utf-8-sig")
    lines = text.splitlines()
    assert lines[0] == "value"
    assert lines[1] == "1"


async def test_download_xlsx(report_client: AsyncClient) -> None:
    resp = await report_client.get(f"{ADMIN_REPORTS}/constant_one/download?format=xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert ".xlsx" in resp.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(resp.content), read_only=True)
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("value",)
    assert rows[1] == (1,)


async def test_download_unknown_report_returns_404(report_client: AsyncClient) -> None:
    resp = await report_client.get(f"{ADMIN_REPORTS}/does_not_exist/download")
    assert resp.status_code == 404


async def test_download_invalid_format_returns_422(report_client: AsyncClient) -> None:
    resp = await report_client.get(f"{ADMIN_REPORTS}/constant_one/download?format=pdf")
    assert resp.status_code == 422


async def test_download_two_columns_csv(report_client: AsyncClient) -> None:
    """Multi-column SELECT round-trips through the CSV writer."""
    resp = await report_client.get(f"{ADMIN_REPORTS}/two_columns/download")
    assert resp.status_code == 200
    lines = resp.content.decode("utf-8-sig").splitlines()
    assert lines[0] == "id,name"
    assert lines[1] == "1,alpha"
